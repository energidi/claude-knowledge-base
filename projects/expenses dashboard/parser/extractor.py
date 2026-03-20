"""
Reads all expense Excel files and returns raw sheet data per file/year.
Handles both .xls (xlrd) and .xlsx (openpyxl) formats.
"""

import os
import re
import openpyxl
import xlrd

EXPENSE_DIR = os.path.join(os.path.dirname(__file__), "..", "Expense Files")

HEBREW_MONTHS = {
    "ינואר": 1, "פברואר": 2, "מרץ": 3, "אפריל": 4,
    "מאי": 5, "יוני": 6, "יולי": 7, "אוגוסט": 8,
    "ספטמבר": 9, "אוקטובר": 10, "נובמבר": 11, "דצמבר": 12,
}

SUMMARY_SHEETS = {"סה\"כ", "סיכום שנה", "סיכום שנתי"}
SKIP_SHEETS = {
    "רשימות", "רשימה", "חשמל", "מים", "תכנון חסכון",
    "השוואה לשנים קודמות", "השוואה ל-2013", "סה\"כ באחוזים",
    "גיליון2", "גיליון3", "עע", "הוצאות טיול בדרום",
    "הוצאות חשמל", "שמות כותרות", "סיכום שכר", "טופס הוצאות והכנסות",
}


def _year_from_filename(filename: str) -> int | None:
    match = re.search(r"(20\d{2}|201\d|200\d)", filename)
    if match:
        return int(match.group(1))
    return None


def _read_xlsx_sheet(wb: openpyxl.Workbook, sheet_name: str) -> list[list]:
    ws = wb[sheet_name]
    return [
        [cell for cell in row]
        for row in ws.iter_rows(values_only=True)
    ]


def _read_xls_sheet(wb: xlrd.Book, sheet_name: str) -> list[list]:
    ws = wb.sheet_by_name(sheet_name)
    return [
        [ws.cell_value(r, c) for c in range(ws.ncols)]
        for r in range(ws.nrows)
    ]


def extract_all() -> dict[int, dict]:
    """
    Returns dict keyed by year, each value:
    {
        "summary": list[list] | None,   # raw rows from summary sheet
        "monthly": { month_num: list[list] }  # raw rows from monthly sheets
    }
    """
    results = {}

    files = [
        f for f in os.listdir(EXPENSE_DIR)
        if f.endswith((".xlsx", ".xls")) and "תבנית" not in f
    ]

    for filename in sorted(files):
        year = _year_from_filename(filename)
        if year is None:
            continue

        path = os.path.join(EXPENSE_DIR, filename)
        is_xlsx = filename.endswith(".xlsx")

        try:
            if is_xlsx:
                wb = openpyxl.load_workbook(path, data_only=True)
                sheet_names = wb.sheetnames
                read_fn = lambda name: _read_xlsx_sheet(wb, name)
            else:
                wb = xlrd.open_workbook(path)
                sheet_names = wb.sheet_names()
                read_fn = lambda name: _read_xls_sheet(wb, name)
        except Exception as e:
            print(f"[extractor] Could not open {filename}: {e}")
            continue

        summary_data = None
        monthly_data = {}

        for name in sheet_names:
            name_stripped = name.strip()
            if name_stripped in SKIP_SHEETS:
                continue
            if name_stripped in SUMMARY_SHEETS:
                try:
                    summary_data = read_fn(name)
                except Exception as e:
                    print(f"[extractor] {filename}/{name}: {e}")
                continue
            month_num = HEBREW_MONTHS.get(name_stripped)
            if month_num:
                try:
                    monthly_data[month_num] = read_fn(name)
                except Exception as e:
                    print(f"[extractor] {filename}/{name}: {e}")

        results[year] = {"summary": summary_data, "monthly": monthly_data}
        print(f"[extractor] {filename} -> year {year}, "
              f"months: {sorted(monthly_data.keys())}, "
              f"summary: {'yes' if summary_data else 'no'}")

    return results
